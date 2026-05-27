import tabula
import pandas as pd
import re
import time
import streamlit as st
import os
import io

st.set_page_config(page_title="PDF to Excel Converter", layout="centered")

st.title("📄 PDF to Excel Converter")
st.write("Drag and drop your PDF below to extract the data into an Excel spreadsheet.")
uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")

if uploaded_file is not None:
    # Save the uploaded file to a temporary location
    pdf_path = "temp.pdf"
    with open(pdf_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
else:
    st.warning("Please upload a PDF file to proceed.")
    st.stop()

with st.spinner("Converting to Excel...", show_time=True):
    SUB_INDEX = "סאב\rאינדקס"
    ARTICLE_NUM = "מס'\rארטיקל"
    PRODUCT_NUM = "מס'\rפרודקט"

    tables = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True)

    indexes_of_interest = ["F 05", "F 40", "F 70", "F 82"]

    df = pd.concat(tables, ignore_index=True)

    # fix df column name in case of typos
    if SUB_INDEX not in df.columns:
        for col in df.columns:
            if SUB_INDEX[:3] in col:
                df.rename(columns={col: SUB_INDEX}, inplace=True)
                break

    all_categories = indexes_of_interest + [x for x in df[SUB_INDEX].unique() if x not in indexes_of_interest]
    df[SUB_INDEX] = pd.Categorical(df[SUB_INDEX], categories=all_categories, ordered=True)
    df = df.sort_values(SUB_INDEX).reset_index(drop=True)

    # df = df[df[SUB_INDEX].isin(indexes_of_interest)]
    df[ARTICLE_NUM] = df[ARTICLE_NUM].apply(lambda x: re.sub("[^0-9]", "", str(x)))

    # make sure PRODUCT_NUM is string and add padding zeros to make sure its of 7 digits
    df[PRODUCT_NUM] = df[PRODUCT_NUM].apply(lambda x: str(x).zfill(7))

    def build_link(row):
        product_num = str(row[PRODUCT_NUM]).strip()
        article_num = str(row[ARTICLE_NUM]).strip()
        return f"https://www2.hm.com/hw_il/productpage.{product_num}{article_num}.html"


    df["link"] = df.apply(build_link, axis=1)

    # Save to Excel with hyperlinks
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active

    # Write dataframe to worksheet
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            # If this is the "מס'\rפרודקט" column (not header), add hyperlink
            if r > 1 and ws.cell(row=1, column=c).value == "מס'\rפרודקט":
                link_value = df.iloc[r-2]['link']
                cell.hyperlink = link_value
                cell.style = "Hyperlink"  # Optional: style as hyperlink

    wb.save('output.xlsx')
    print("Excel file 'output.xlsx' created with hyperlinks in the product number column.")

st.success("Conversion to Excel complete!")


# 4. The Download Button
st.download_button(
    label="📥 Download Excel File",
    data=open('output.xlsx', 'rb').read(),
    file_name="converted_data.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

with st.spinner("Converting to PDF...", show_time=True):
    # convert df link to html hyperlink
    df["link"] = df["link"].apply(lambda x: f'<a href="{x}">{x}</a>')
    # convert PRODUCT_NUM to same hyperlink but with the product number as the text
    df[PRODUCT_NUM] = df.apply(lambda row: f'<a href="{row["link"]}">{row[PRODUCT_NUM]}</a>', axis=1)

    html_table = df.to_html(index=False, escape=False)

    from weasyprint import HTML

    html_string = f"""
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; font-size: 10px; }}
        table {{ border-collapse: collapse; width: 100%; font-size: 10px; }}
        th, td {{ padding: 3px; text-align: middle; border: 1px solid #ddd; }}
        th {{ background-color: #f2f2f2; }}
    </style>
</head>
<body>
    {html_table}
</body>
</html>
"""

    pdf_buffer = io.BytesIO()
    HTML(string=html_string).write_pdf(pdf_buffer)

    # use ILOVEPDF API to convert the excel file to pdf

    # from iloveapi import ILoveApi

    # client = ILoveApi(
    #     public_key=os.getenv("ILOVEPDF_PUBLIC_KEY"),
    #     secret_key=os.getenv("ILOVEPDF_SECRET_KEY"),
    # )
    # task = client.create_task("officepdf")
    # task.add_file("output.xlsx")
    # task.process()
    # task.download("output.pdf")


st.success("Conversion to PDF complete!")


# 4. The Download Button
st.download_button(
    label="📥 Download PDF File",
    data=pdf_buffer.getvalue(),
    file_name="converted_data.pdf",
    mime="application/pdf"
)

# 5. Restart Button
if st.button("🔄 Start Again"):
    uploaded_file = None
    st.rerun()

